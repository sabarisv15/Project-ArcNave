#!/usr/bin/env python3
"""Recalculate a generated workbook and report what its formulas actually did.

This is the quality gate for sandbox-produced .xlsx files. It is based on
the consumer xlsx skill's own recalc.py, but is deliberately stricter in
two ways that script does not cover.

WHY A ZERO EXIT CODE FROM LIBREOFFICE IS NOT A GATE
---------------------------------------------------
`soffice --convert-to xlsx` succeeding proves only that LibreOffice could
open and re-save the file. It says nothing about whether the generated
formulas mean anything. Three things can be wrong in a workbook that
converts perfectly:

  1. A formula evaluates to an error. The real case this exists for: a
     merged month-header cell left a SUMIFS pointing at a blank column,
     and an entire Credit column silently came out as zero. Text
     inspection of the file would not have caught it.

  2. A formula was never a formula. A generator that computes the total
     in Python and writes `170722.00` produces a workbook where every
     number is CORRECT and which is still not what was asked for: edit a
     source row and the total does not move. Nothing else detects this,
     which is why `expect_formulas_in` is required rather than optional.

  3. A formula has no cached value after recalculation. openpyxl writes
     formulas with no cached values; if LibreOffice did not evaluate one,
     `data_only=True` returns None. That is neither an error nor a pass,
     and folding it into "0 errors" is exactly how a gate stops meaning
     anything.

Each is reported separately. The caller decides; this script does not
collapse them into a boolean.

USAGE
-----
    recalc.py <workbook.xlsx> [expect.json]

`expect.json` is a list of cell references or ranges, optionally
sheet-qualified:

    ["Summary!B2:B9", "Summary!D2", "C10:C20"]

Writes a JSON report to stdout. Exit code is 0 whenever the CHECK RAN —
a failing workbook still exits 0 with `"passed": false`, because a
non-zero exit cannot be distinguished from the script itself crashing.
Exit 1 means the check could not be performed at all.
"""

import json
import os
import subprocess
import sys
import tempfile

try:
    import openpyxl
    from openpyxl.utils import range_boundaries, get_column_letter
except ImportError:  # pragma: no cover - image build guarantees this
    print(json.dumps({"error": "openpyxl is not available in this image"}))
    sys.exit(1)


# Excel's error strings, as openpyxl surfaces them from a cached value.
ERROR_VALUES = {
    "#REF!", "#VALUE!", "#NAME?", "#DIV/0!", "#N/A", "#NULL!", "#NUM!",
}

LIBREOFFICE_TIMEOUT_SECONDS = 120


def recalculate(path, workdir):
    """Round-trip the workbook through LibreOffice so formulas get evaluated.

    Returns the path to the recalculated copy. LibreOffice is told to use
    a scratch profile inside workdir: a shared default profile is process-
    global state, and this service runs a fresh child per request
    precisely so that no such state carries between executions.
    """
    outdir = os.path.join(workdir, "recalculated")
    os.makedirs(outdir, exist_ok=True)
    profile = os.path.join(workdir, "lo-profile")
    result = subprocess.run(
        [
            "soffice",
            "--headless",
            "--norestore",
            "--invisible",
            f"-env:UserInstallation=file://{profile}",
            "--convert-to", "xlsx",
            "--outdir", outdir,
            path,
        ],
        capture_output=True,
        timeout=LIBREOFFICE_TIMEOUT_SECONDS,
    )
    produced = os.path.join(outdir, os.path.basename(path))
    if not os.path.exists(produced):
        raise RuntimeError(
            "LibreOffice produced no output file "
            f"(exit {result.returncode}): {result.stderr.decode('utf-8', 'replace')[:500]}"
        )
    return produced


def expand(expectations, default_sheet):
    """Turn ["Summary!B2:B9", "C10"] into {(sheet, "B2"), ...}."""
    cells = set()
    for raw in expectations:
        ref = str(raw).strip()
        if not ref:
            continue
        if "!" in ref:
            sheet, _, ref = ref.partition("!")
            sheet = sheet.strip().strip("'")
        else:
            sheet = default_sheet
        min_col, min_row, max_col, max_row = range_boundaries(ref)
        for row in range(min_row, max_row + 1):
            for col in range(min_col, max_col + 1):
                cells.add((sheet, f"{get_column_letter(col)}{row}"))
    return cells


def inspect(path, expectations):
    # Two loads of the SAME recalculated file: one reports what each cell
    # still says, the other what it evaluated to. Both are needed —
    # neither alone can distinguish the three failure modes above.
    formulas_wb = openpyxl.load_workbook(path, data_only=False)
    values_wb = openpyxl.load_workbook(path, data_only=True)

    default_sheet = formulas_wb.sheetnames[0]
    expected_cells = expand(expectations, default_sheet)

    error_cells = []
    uncached_cells = []
    formula_cells = []

    for sheet_name in formulas_wb.sheetnames:
        formula_sheet = formulas_wb[sheet_name]
        value_sheet = values_wb[sheet_name]
        for row in formula_sheet.iter_rows():
            for cell in row:
                raw = cell.value
                if not (isinstance(raw, str) and raw.startswith("=")):
                    continue
                formula_cells.append(f"{sheet_name}!{cell.coordinate}")
                cached = value_sheet[cell.coordinate].value
                if isinstance(cached, str) and cached.strip() in ERROR_VALUES:
                    error_cells.append({
                        "cell": f"{sheet_name}!{cell.coordinate}",
                        "formula": raw,
                        "value": cached.strip(),
                    })
                elif cached is None:
                    uncached_cells.append({
                        "cell": f"{sheet_name}!{cell.coordinate}",
                        "formula": raw,
                    })

    # Failure mode 2: a cell the caller declared as a formula cell that
    # holds something else. A plain number here is the dangerous case —
    # the workbook looks right and is dead.
    constant_cells = []
    missing_cells = []
    for sheet_name, coordinate in sorted(expected_cells):
        if sheet_name not in formulas_wb.sheetnames:
            missing_cells.append(f"{sheet_name}!{coordinate}")
            continue
        raw = formulas_wb[sheet_name][coordinate].value
        if isinstance(raw, str) and raw.startswith("="):
            continue
        constant_cells.append({
            "cell": f"{sheet_name}!{coordinate}",
            "found": raw,
            "reason": "expected a formula, found a literal value — a total computed "
                      "in code and written as a number does not update when a source "
                      "row changes",
        })

    return {
        "formulaCellCount": len(formula_cells),
        "expectedFormulaCellCount": len(expected_cells),
        "errors": error_cells,
        "uncached": uncached_cells,
        "constants": constant_cells,
        "missingCells": missing_cells,
    }


def main():
    if len(sys.argv) < 2:
        print(json.dumps({"error": "usage: recalc.py <workbook.xlsx> [expect.json]"}))
        return 1

    workbook = sys.argv[1]
    if not os.path.exists(workbook):
        print(json.dumps({"error": f"no such workbook: {workbook}"}))
        return 1

    expectations = []
    if len(sys.argv) > 2 and os.path.exists(sys.argv[2]):
        with open(sys.argv[2], "r", encoding="utf-8") as handle:
            expectations = json.load(handle)
    if not isinstance(expectations, list):
        print(json.dumps({"error": "expect.json must be a JSON array of cell references"}))
        return 1

    with tempfile.TemporaryDirectory() as workdir:
        try:
            recalculated = recalculate(workbook, workdir)
        except Exception as err:  # noqa: BLE001 - reported, not swallowed
            print(json.dumps({
                "verdict": "unverified",
                "passed": False,
                "reason": "recalculation_failed",
                "detail": str(err)[:500],
            }))
            return 0
        report = inspect(recalculated, expectations)

    # An undeclared workbook is UNVERIFIED, never "passed". Without a
    # declared expectation, failure mode 2 has no referent — there is no
    # way to know which cells were supposed to be formulas, so the check
    # genuinely did not happen and must not be reported as if it had.
    if not expectations:
        verdict = "unverified"
        reason = "no expect_formulas_in was declared, so formula cells could not be checked"
    elif report["errors"] or report["constants"] or report["uncached"] or report["missingCells"]:
        verdict = "failed"
        reason = "; ".join(filter(None, [
            f"{len(report['errors'])} formula error(s)" if report["errors"] else "",
            f"{len(report['constants'])} expected formula cell(s) hold literal values" if report["constants"] else "",
            f"{len(report['uncached'])} formula cell(s) were not evaluated" if report["uncached"] else "",
            f"{len(report['missingCells'])} expected cell(s) do not exist" if report["missingCells"] else "",
        ]))
    else:
        verdict = "passed"
        reason = (
            f"{report['formulaCellCount']} formula cell(s) recalculated cleanly; "
            f"{report['expectedFormulaCellCount']} declared cell(s) still hold formulas"
        )

    print(json.dumps({
        "verdict": verdict,
        "passed": verdict == "passed",
        "reason": reason,
        **report,
    }))
    return 0


if __name__ == "__main__":
    sys.exit(main())
