#!/usr/bin/env python3
"""Tests for recalc.py — the xlsx quality gate.

Two tiers, deliberately separated:

  - PURE tests need only Python + openpyxl. They run anywhere, including
    outside the sandbox image, and cover expand() and the "uncached"
    failure mode (which does not need LibreOffice at all — an
    openpyxl-written file that was NEVER recalculated has no cached
    values by construction, which is the exact condition "uncached"
    detects).

  - LIBREOFFICE tests actually shell out to `soffice` and prove the full
    chain end to end: this is what confirms LibreOffice's headless
    --convert-to genuinely recalculates and caches formula results
    (verified manually against this exact image before this file was
    written — SUMIF, #DIV/0!, and an unresolved external reference all
    came back with real cached values, not None). These are skipped with
    an explicit message when `soffice` is not on PATH, never silently
    treated as passing.

Run: python3 scripts/test_recalc.py
"""

import json
import os
import shutil
import sys
import tempfile
import unittest

sys.path.insert(0, os.path.dirname(__file__))
import recalc  # noqa: E402

import openpyxl  # noqa: E402


class ExpandTests(unittest.TestCase):
    def test_single_cell_defaults_to_first_sheet(self):
        cells = recalc.expand(["B2"], "Summary")
        self.assertEqual(cells, {("Summary", "B2")})

    def test_range_expands_every_cell(self):
        cells = recalc.expand(["B2:B3"], "Summary")
        self.assertEqual(cells, {("Summary", "B2"), ("Summary", "B3")})

    def test_sheet_qualified_reference_overrides_default(self):
        cells = recalc.expand(["Totals!C1"], "Summary")
        self.assertEqual(cells, {("Totals", "C1")})

    def test_quoted_sheet_name(self):
        cells = recalc.expand(["'My Sheet'!A1"], "Summary")
        self.assertEqual(cells, {("My Sheet", "A1")})

    def test_empty_and_blank_entries_are_ignored(self):
        cells = recalc.expand(["", "  ", "B2"], "Summary")
        self.assertEqual(cells, {("Summary", "B2")})


class UncachedDetectionTests(unittest.TestCase):
    """No LibreOffice needed: an openpyxl-written file was never
    recalculated, so every formula cell has no cached value by
    construction. This is the real, reachable path for the 'uncached'
    verdict — not a synthetic scenario."""

    def test_formula_cell_with_no_cached_value_is_reported_uncached(self):
        with tempfile.TemporaryDirectory() as tmp:
            path = os.path.join(tmp, "book.xlsx")
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Summary"
            ws["A1"] = 10
            ws["B1"] = "=A1*2"
            wb.save(path)

            report = recalc.inspect(path, ["Summary!B1"])

            self.assertEqual(report["uncached"], [{"cell": "Summary!B1", "formula": "=A1*2"}])
            self.assertEqual(report["errors"], [])
            self.assertEqual(report["constants"], [])

    def test_non_formula_cells_are_never_inspected(self):
        with tempfile.TemporaryDirectory() as tmp:
            path = os.path.join(tmp, "book.xlsx")
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Summary"
            ws["A1"] = 10
            ws["A2"] = "just text"
            wb.save(path)

            report = recalc.inspect(path, [])

            self.assertEqual(report["formulaCellCount"], 0)


@unittest.skipUnless(shutil.which("soffice"), "soffice not on PATH — these tests only run inside the sandbox image")
class LibreOfficeRoundTripTests(unittest.TestCase):
    """Full pipeline: build with openpyxl, recalculate with LibreOffice,
    inspect the result. Each test is the fixture the approved plan
    required, one per distinct failure mode."""

    def run_gate(self, build_workbook, expect_formulas_in):
        with tempfile.TemporaryDirectory() as tmp:
            path = os.path.join(tmp, "book.xlsx")
            build_workbook(path)
            recalculated = recalc.recalculate(path, tmp)
            return recalc.inspect(recalculated, expect_formulas_in)

    def test_pass_sumif_over_valid_range(self):
        def build(path):
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Summary"
            ws["A1"], ws["B1"] = "Category", "Amount"
            ws["A2"], ws["B2"] = "PLB", 100
            ws["A3"], ws["B3"] = "PLB", 200
            ws["A4"], ws["B4"] = "SD", 50
            ws["D2"], ws["E2"] = "PLB", "=SUMIF(A2:A4,D2,B2:B4)"
            wb.save(path)

        report = self.run_gate(build, ["Summary!E2"])
        self.assertEqual(report["errors"], [])
        self.assertEqual(report["constants"], [])
        self.assertEqual(report["uncached"], [])

    def test_error_value_from_a_formula_pointing_at_a_bad_reference(self):
        # Stands in for the real bug this gate exists for: a formula
        # left pointing at a blank/merged cell, caught as a cached
        # error value after recalculation.
        def build(path):
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Summary"
            ws["A1"] = 10
            ws["B1"] = "=A1/0"
            wb.save(path)

        report = self.run_gate(build, ["Summary!B1"])
        self.assertEqual(len(report["errors"]), 1)
        self.assertEqual(report["errors"][0]["value"], "#DIV/0!")

    def test_expected_formula_cell_holding_a_literal_constant_fails(self):
        # THE distinguishing fixture: every number in this workbook is
        # correct. It must still fail, because "0 errors" from
        # LibreOffice says nothing about whether a formula was ever
        # there — a generator that computes the answer in Python and
        # writes it as a number produces a workbook that looks perfect
        # and does not update when a source row changes.
        def build(path):
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Summary"
            ws["A1"] = 100
            ws["A2"] = 200
            ws["B1"] = 300  # correct value, not a formula
            wb.save(path)

        report = self.run_gate(build, ["Summary!B1"])
        self.assertEqual(report["errors"], [])
        self.assertEqual(len(report["constants"]), 1)
        self.assertEqual(report["constants"][0]["found"], 300)

    def test_undeclared_workbook_cannot_be_reported_as_passed(self):
        # main()'s own verdict logic (not inspect() in isolation) is what
        # enforces "unverified, never passed" when nothing was declared —
        # covered directly against the CLI entry point here.
        with tempfile.TemporaryDirectory() as tmp:
            path = os.path.join(tmp, "book.xlsx")
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Summary"
            ws["A1"] = "=1+1"
            wb.save(path)

            old_argv = sys.argv
            old_stdout = sys.stdout
            sys.argv = ["recalc.py", path]
            sys.stdout = captured = __import__("io").StringIO()
            try:
                recalc.main()
            finally:
                sys.argv = old_argv
                sys.stdout = old_stdout

            result = json.loads(captured.getvalue())
            self.assertEqual(result["verdict"], "unverified")
            self.assertFalse(result["passed"])


if __name__ == "__main__":
    unittest.main()
